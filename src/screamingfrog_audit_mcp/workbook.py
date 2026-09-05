"""The branded master workbook.

Every export the crawl produced becomes its own sheet, styled the same way, so
a reader can pick up any tab and know how to read it: purple header row,
banded rows, frozen header, autofilter, sized columns, coloured tab.

Highlighting is the part that makes it an audit rather than a data dump.
Cells are tinted where the value carries meaning: issue priority, HTTP status
class, indexability, thin content, slow responses, missing titles and H1s.
A reader should see the problems before reading a single cell.
"""

from __future__ import annotations

import csv
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from . import branding as B

FONT = "Aptos Narrow"
MAX_ROWS_PER_SHEET = 5000      # keeps the file openable on a laptop
MAX_COL_WIDTH = 58

# Sheets whose rows carry a status code worth tinting.
_STATUS_COLUMNS = ("Status Code",)
_THIN_WORDS = 300
_SLOW_SECONDS = 1.0


def _fill(hexcode: str) -> PatternFill:
    return PatternFill("solid", fgColor=hexcode)


def _num(value, default=None):
    try:
        return float(str(value).replace(",", "").strip())
    except (TypeError, ValueError, AttributeError):
        return default


def _safe_sheet_name(name: str, taken: set[str]) -> str:
    """Excel: 31 chars, and none of []:*?/\\ ."""
    clean = name.replace(".csv", "")
    for ch in "[]:*?/\\":
        clean = clean.replace(ch, "-")
    clean = clean.replace("_", " ").strip()[:31] or "Sheet"
    base, n = clean, 2
    while clean.lower() in taken:
        suffix = f" {n}"
        clean = base[: 31 - len(suffix)] + suffix
        n += 1
    taken.add(clean.lower())
    return clean


class BrandedWorkbook:
    def __init__(self):
        self.wb = Workbook()
        self.wb.remove(self.wb.active)
        self._names: set[str] = set()
        # export filename -> what its sheet actually holds. consolidate() will
        # not delete a CSV unless its entry here says the sheet is complete.
        self.carried: dict[str, dict] = {}

        self.hdr_fill = _fill(B.PURPLE)
        self.hdr_font = Font(color=B.WHITE, bold=True, size=10, name=FONT)
        self.band_fill = _fill(B.LIGHT_CREAM)
        self.cream_fill = _fill(B.CREAM)
        self.sand_fill = _fill(B.CARD_SAND)
        self.title_font = Font(bold=True, size=22, color=B.PURPLE, name=FONT)
        self.h2_font = Font(bold=True, size=12, color=B.TEXT, name=FONT)
        self.body_font = Font(size=10, color=B.TEXT_2, name=FONT)
        self.muted_font = Font(size=9, color=B.MUTED, name=FONT, italic=True)
        thin = Side(style="thin", color=B.BORDER)
        self.box = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── shared styling ───────────────────────────────────────────────────────

    def _style_table(self, ws, header_row: int, ncols: int, nrows: int):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=header_row, column=c)
            cell.fill, cell.font = self.hdr_fill, self.hdr_font
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        ws.row_dimensions[header_row].height = 28

        for r in range(header_row + 1, header_row + nrows + 1):
            for c in range(1, ncols + 1):
                cell = ws.cell(row=r, column=c)
                cell.font = self.body_font
                cell.border = self.box
                cell.alignment = Alignment(vertical="top")
                if (r - header_row) % 2 == 0:
                    cell.fill = self.band_fill

        ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
        if nrows:
            ws.auto_filter.ref = (
                f"A{header_row}:{get_column_letter(ncols)}{header_row + nrows}")

    def _autosize(self, ws, ncols: int):
        for c in range(1, ncols + 1):
            widest = 10
            for cell in ws[get_column_letter(c)]:
                if cell.value is not None:
                    widest = max(widest, min(len(str(cell.value)) + 2, MAX_COL_WIDTH))
            ws.column_dimensions[get_column_letter(c)].width = widest

    def _brand_header(self, ws, title: str, subtitle: str = ""):
        ws["A1"] = title
        ws["A1"].font = self.title_font
        ws.row_dimensions[1].height = 30
        if subtitle:
            ws["A2"] = subtitle
            ws["A2"].font = self.body_font
        ws["A3"] = B.CREDIT_LINE
        ws["A3"].font = self.muted_font

    def sheet(self, name: str, tab_colour: str = B.PURPLE):
        ws = self.wb.create_sheet(_safe_sheet_name(name, self._names))
        ws.sheet_properties.tabColor = tab_colour
        ws.sheet_view.showGridLines = False
        return ws

    # ── the sheets ───────────────────────────────────────────────────────────

    def add_summary(self, summary: dict, site: str, exports: int):
        ws = self.sheet("Summary", B.PURPLE)
        stats = summary.get("stats", {})
        counts = summary.get("counts", {})
        health = summary.get("health", {})
        score = health.get("score", 0)
        label, colour = B.band(score)

        self._brand_header(ws, B.report_title(site),
                           f"Crawled {summary.get('crawled_at', '')}")

        ws["A5"] = "Health score"
        ws["A5"].font = self.h2_font
        ws["B5"] = score
        ws["B5"].font = Font(bold=True, size=36, color=colour, name=FONT)
        ws["C5"] = label
        ws["C5"].font = Font(bold=True, size=12, color=colour, name=FONT)
        ws["A6"] = health.get("formula", "")
        ws["A6"].font = self.muted_font

        kpis = [
            ("URLs crawled", stats.get("urls", 0), B.TEXT),
            ("Indexable", stats.get("indexable", 0), B.GREEN),
            ("Non-indexable", stats.get("non_indexable", 0), B.MUTED),
            ("High priority issues", counts.get("high", 0), B.RED),
            ("Medium priority issues", counts.get("medium", 0), B.AMBER),
            ("Low priority issues", counts.get("low", 0), B.MUTED),
            ("Issue types found", counts.get("total_types", 0), B.TEXT),
            ("Data tables in this workbook", exports, B.TEXT),
        ]
        row = 8
        ws.cell(row=row, column=1, value="Headline").font = self.h2_font
        row += 1
        for name, value, colour_ in kpis:
            ws.cell(row=row, column=1, value=name).font = self.body_font
            c = ws.cell(row=row, column=2, value=value)
            c.font = Font(bold=True, size=12, color=colour_, name=FONT)
            for col in (1, 2):
                ws.cell(row=row, column=col).fill = self.sand_fill
                ws.cell(row=row, column=col).border = self.box
            row += 1

        if stats.get("status"):
            row += 1
            ws.cell(row=row, column=1, value="Status codes").font = self.h2_font
            row += 1
            for code, n in sorted(stats["status"].items()):
                ws.cell(row=row, column=1, value=code).font = self.body_font
                ws.cell(row=row, column=2, value=n).font = self.body_font
                tint = B.STATUS_TINT.get(str(code)[:1])
                for col in (1, 2):
                    cell = ws.cell(row=row, column=col)
                    cell.border = self.box
                    if tint:
                        cell.fill = _fill(tint)
                row += 1

        row += 2
        ws.cell(row=row, column=1, value=B.CREDIT_LONG).font = self.muted_font

        ws.column_dimensions["A"].width = 34
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 18
        return ws

    def add_issues(self, issues: list[dict]):
        ws = self.sheet("Issue register", B.RED if issues else B.GREEN)
        self._brand_header(ws, "Issue register",
                           "Sorted by priority, then by URLs affected")

        headers = ["Priority", "Type", "URLs", "% of crawl", "Issue", "How to fix"]
        header_row = 5
        for i, h in enumerate(headers, 1):
            ws.cell(row=header_row, column=i, value=h)

        for r, issue in enumerate(issues, header_row + 1):
            ws.cell(row=r, column=1, value=issue.get("priority"))
            ws.cell(row=r, column=2, value=issue.get("type"))
            ws.cell(row=r, column=3, value=issue.get("urls"))
            ws.cell(row=r, column=4, value=issue.get("pct"))
            ws.cell(row=r, column=5, value=issue.get("issue"))
            ws.cell(row=r, column=6, value=issue.get("how_to_fix"))

        self._style_table(ws, header_row, len(headers), len(issues))

        # The highlight that matters: priority reads as colour, not as a word.
        for r, issue in enumerate(issues, header_row + 1):
            pri = issue.get("priority", "")
            chip = ws.cell(row=r, column=1)
            chip.fill = _fill(B.PRIORITY_FILL.get(pri, B.MUTED))
            chip.font = Font(bold=True, size=10, color=B.WHITE, name=FONT)
            chip.alignment = Alignment(horizontal="center", vertical="center")
            tint = B.PRIORITY_TINT.get(pri)
            if tint:
                for c in range(2, len(headers) + 1):
                    ws.cell(row=r, column=c).fill = _fill(tint)
            ws.cell(row=r, column=6).alignment = Alignment(vertical="top", wrap_text=True)

        for col, width in zip("ABCDEF", (12, 13, 9, 11, 46, 70)):
            ws.column_dimensions[col].width = width
        return ws

    def add_analysis(self, analysis: dict):
        """The derived layer: what the set of URLs means."""
        ws = self.sheet("Analysis", B.VIOLET)
        self._brand_header(ws, "What the set of URLs means",
                           "Derived from the crawl, not exported from it")

        row = 5
        for key in ("depth", "link_equity", "sitemap", "content",
                    "performance", "indexability", "duplication"):
            section = analysis.get(key)
            if not isinstance(section, dict):
                continue
            ws.cell(row=row, column=1,
                    value=key.replace("_", " ").title()).font = self.h2_font
            ws.cell(row=row, column=1).fill = self.sand_fill
            row += 1
            for k, v in section.items():
                if k == "reading" or isinstance(v, (list, dict)):
                    continue
                ws.cell(row=row, column=1, value=k.replace("_", " ")).font = self.body_font
                c = ws.cell(row=row, column=2, value=v)
                c.font = Font(bold=True, size=10, color=B.TEXT, name=FONT)
                for col in (1, 2):
                    ws.cell(row=row, column=col).border = self.box
                row += 1
            reading = section.get("reading")
            if reading:
                cell = ws.cell(row=row, column=1, value=reading)
                cell.font = self.muted_font
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
                ws.row_dimensions[row].height = 30
                row += 1
            row += 1

        ws.column_dimensions["A"].width = 38
        ws.column_dimensions["B"].width = 20
        for col in "CD":
            ws.column_dimensions[col].width = 24
        return ws

    def add_export(self, path: Path):
        """One crawl export, styled and highlighted."""
        with open(path, newline="", encoding="utf-8-sig") as fh:
            reader = csv.reader(fh)
            try:
                headers = next(reader)
            except StopIteration:
                return None
            rows, total = [], 0
            for row in reader:
                total += 1
                if len(rows) < MAX_ROWS_PER_SHEET:
                    rows.append(row)

        if not rows:
            return None
        truncated = total > len(rows)

        ws = self.sheet(path.stem, B.CARD_SAND)
        for i, h in enumerate(headers, 1):
            ws.cell(row=1, column=i, value=h)
        for r, row in enumerate(rows, 2):
            for c, value in enumerate(row[: len(headers)], 1):
                ws.cell(row=r, column=c, value=value)

        self._style_table(ws, 1, len(headers), len(rows))
        self._highlight(ws, headers, len(rows))
        self._autosize(ws, len(headers))
        # Recorded so the caller can tell a complete sheet from a sampled one.
        # Only a complete sheet may stand in for its CSV.
        self.carried[path.name] = {
            "sheet": ws.title, "rows_written": len(rows), "rows_total": total,
            "columns": len(headers), "complete": not truncated,
        }

        if truncated:
            note = ws.cell(row=len(rows) + 3, column=1,
                           value=f"Showing the first {MAX_ROWS_PER_SHEET} of "
                                 f"{total} rows. Full data: {path.name}")
            note.font = self.muted_font
        return ws

    def _highlight(self, ws, headers: list[str], nrows: int):
        """Tint the cells whose value is the finding."""
        idx = {h: i + 1 for i, h in enumerate(headers)}

        def col(*names):
            for n in names:
                if n in idx:
                    return idx[n]
            return None

        status_c = col(*_STATUS_COLUMNS)
        index_c = col("Indexability")
        words_c = col("Word Count")
        resp_c = col("Response Time")

        for r in range(2, nrows + 2):
            if status_c:
                code = str(ws.cell(row=r, column=status_c).value or "")
                tint = B.STATUS_TINT.get(code[:1])
                if tint and code[:1] not in ("2",):
                    ws.cell(row=r, column=status_c).fill = _fill(tint)
                    ws.cell(row=r, column=status_c).font = Font(
                        bold=True, size=10, name=FONT,
                        color=B.RED if code[:1] in ("4", "5") else B.TEXT_2)
            if index_c:
                cell = ws.cell(row=r, column=index_c)
                if str(cell.value or "").strip().lower().startswith("non"):
                    cell.fill = _fill(B.PRIORITY_TINT["Low"])
                    cell.font = Font(size=10, color=B.MUTED, name=FONT)
            if words_c:
                cell = ws.cell(row=r, column=words_c)
                n = _num(cell.value)
                if n is not None and 0 < n < _THIN_WORDS:
                    cell.fill = _fill(B.PRIORITY_TINT["Medium"])
                    cell.font = Font(bold=True, size=10, color=B.AMBER, name=FONT)
            if resp_c:
                cell = ws.cell(row=r, column=resp_c)
                n = _num(cell.value)
                if n is not None and n >= _SLOW_SECONDS:
                    cell.fill = _fill(B.PRIORITY_TINT["High"])
                    cell.font = Font(bold=True, size=10, color=B.RED, name=FONT)

    def add_index(self, entries: list[tuple[str, int]]):
        ws = self.sheet("Data index", B.MUTED)
        self._brand_header(ws, "Data index",
                           "Every table in this workbook, largest first")
        header_row = 5
        ws.cell(row=header_row, column=1, value="Sheet")
        ws.cell(row=header_row, column=2, value="Rows")
        for r, (name, n) in enumerate(entries, header_row + 1):
            ws.cell(row=r, column=1, value=name)
            ws.cell(row=r, column=2, value=n)
        self._style_table(ws, header_row, 2, len(entries))
        ws.column_dimensions["A"].width = 40
        ws.column_dimensions["B"].width = 12
        return ws

    def save(self, path: Path):
        self.wb.save(path)
        return path


def build(folder: Path, summary: dict, analysis: dict, site: str,
          out_path: Path) -> dict:
    """Assemble the whole workbook for a finished crawl."""
    exports = sorted(folder.glob("*.csv"))
    bw = BrandedWorkbook()

    bw.add_summary(summary, site, len(exports))
    bw.add_issues(summary.get("issues", []))
    bw.add_analysis(analysis)

    written: list[tuple[str, int]] = []
    for path in exports:
        ws = bw.add_export(path)
        if ws is not None:
            written.append((ws.title, ws.max_row - 1))

    written.sort(key=lambda x: -x[1])
    bw.add_index(written)

    bw.save(out_path)

    # An export with a header and no data rows gets no sheet, and loses nothing
    # by being removed, so it counts as carried.
    carried = dict(bw.carried)
    for path in exports:
        carried.setdefault(path.name, {
            "sheet": None, "rows_written": 0, "rows_total": 0,
            "columns": 0, "complete": True,
        })

    return {
        "path": str(out_path),
        "sheets": len(bw.wb.sheetnames),
        "data_tables": len(written),
        "carried": carried,
        "generated_at": datetime.now().isoformat(timespec="seconds"),
    }
