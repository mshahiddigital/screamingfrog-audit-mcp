"""Fold a crawl's CSV exports into the master workbook, then remove them.

A finished crawl folder holds seventy-odd CSV files that nobody opens
individually. The workbook already carries every one of them as its own sheet,
so once that is true the CSVs are duplication, not a safety net.

The invariant is the whole point: a file is deleted only after its sheet has
been read back off disk and confirmed to hold every row. Anything the workbook
had to sample or could not carry stays on disk and is named in the manifest, so
a consolidated folder is never quietly incomplete.

This is opt-in. read_export and aggregate_export need the CSVs, so a crawl you
are still asking questions about should keep them. Consolidate when the folder
is a finished deliverable.
"""

from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path

MANIFEST = "consolidated.json"


def is_consolidated(folder: Path) -> bool:
    """True when this folder's exports were folded into the workbook."""
    return (folder / MANIFEST).exists() and not any(folder.glob("*.csv"))


def manifest(folder: Path) -> dict:
    path = folder / MANIFEST
    if not path.exists():
        return {}
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return {}


def verify(workbook_path: Path, carried: dict) -> tuple[bool, str]:
    """Reopen the saved workbook and confirm each sheet holds its rows.

    Deleting on the strength of "we think we wrote it" is how data disappears.
    A truncated save or a dropped sheet stops the cleanup here instead.
    """
    if not workbook_path.exists() or workbook_path.stat().st_size == 0:
        return False, "no workbook was written"
    try:
        from openpyxl import load_workbook
        wb = load_workbook(workbook_path, read_only=True)
    except Exception as exc:                      # noqa: BLE001 - reported, not raised
        return False, f"could not reopen the workbook ({exc})"
    try:
        names = set(wb.sheetnames)
        for filename, rec in carried.items():
            sheet = rec.get("sheet")
            if sheet is None:
                continue
            if sheet not in names:
                return False, f"sheet {sheet!r} for {filename} is missing"
            got = (wb[sheet].max_row or 0) - 1     # the header costs one row
            if got < rec["rows_written"]:
                return False, (f"sheet {sheet!r} holds {got} rows, "
                               f"expected {rec['rows_written']}")
    finally:
        wb.close()
    return True, "verified"


def run(folder: Path, workbook_path: Path, carried: dict) -> dict:
    """Verify the workbook, then delete every export it carries in full."""
    ok, detail = verify(workbook_path, carried)
    if not ok:
        return {"ok": False, "reason": f"workbook verification failed: {detail}",
                "deleted_count": 0, "deleted": [], "kept": []}

    deleted, kept = [], []
    for path in sorted(folder.glob("*.csv")):
        rec = carried.get(path.name)
        if rec is None:
            kept.append({"file": path.name, "reason": "not carried by the workbook"})
            continue
        if not rec["complete"]:
            kept.append({"file": path.name,
                         "reason": f"sheet holds the first {rec['rows_written']} of "
                                   f"{rec['rows_total']} rows"})
            continue
        try:
            path.unlink()
        except OSError as exc:
            kept.append({"file": path.name, "reason": f"could not delete ({exc})"})
            continue
        deleted.append({"file": path.name, "rows": rec["rows_total"],
                        "columns": rec["columns"],
                        "sheet": rec["sheet"] or "empty, nothing to carry"})

    result = {
        "consolidated_at": datetime.now().isoformat(timespec="seconds"),
        "workbook": workbook_path.name,
        "deleted_count": len(deleted),
        "kept_count": len(kept),
        "note": ("Each export below was written into the workbook in full, "
                 "verified by reopening the saved file, and then removed. The "
                 "workbook is the single source for this crawl's data. "
                 "read_export and aggregate_export need the CSVs, so crawl "
                 "again if this folder has to answer questions."),
        "deleted": deleted,
        "kept": kept,
    }
    (folder / MANIFEST).write_text(json.dumps(result, indent=2), encoding="utf-8")
    result["ok"] = True
    return result
