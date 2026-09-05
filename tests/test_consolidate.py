"""Consolidation must never delete a CSV the workbook does not fully carry.

These tests exist because the failure mode is silent and unrecoverable: a
trimmed sheet plus a deleted export looks exactly like a clean folder.
"""

import csv
from pathlib import Path

import pytest

from screamingfrog_audit_mcp import consolidate, report, workbook

pytest.importorskip("openpyxl")

from test_analysis import crawl  # noqa: F401,E402  (the shared fixture)


def _csv_names(folder: Path) -> set[str]:
    return {p.name for p in folder.glob("*.csv")}


def test_consolidate_removes_every_carried_export(crawl: Path):  # noqa: F811
    before = _csv_names(crawl)
    assert before, "the fixture should start with exports on disk"

    result = report.build(crawl, "Example", consolidate=True)["consolidation"]

    assert result["ok"]
    assert result["kept"] == []
    assert {d["file"] for d in result["deleted"]} == before
    assert _csv_names(crawl) == set()
    assert (crawl / "audit-workbook.xlsx").exists()
    assert consolidate.is_consolidated(crawl)


def test_every_deleted_row_survives_in_the_workbook(crawl: Path):  # noqa: F811
    from openpyxl import load_workbook

    originals = {}
    for path in crawl.glob("*.csv"):
        with open(path, newline="", encoding="utf-8-sig") as fh:
            originals[path.name] = list(csv.DictReader(fh))

    report.build(crawl, "Example", consolidate=True)
    manifest = consolidate.manifest(crawl)
    wb = load_workbook(crawl / "audit-workbook.xlsx", read_only=True)
    try:
        for record in manifest["deleted"]:
            rows = originals[record["file"]]
            if not rows:
                continue
            ws = wb[record["sheet"]]
            assert (ws.max_row or 0) - 1 == len(rows), record["file"]
            assert (ws.max_column or 0) == len(rows[0]), record["file"]
    finally:
        wb.close()


def test_a_sampled_sheet_keeps_its_csv(crawl: Path, monkeypatch):  # noqa: F811
    """A sheet holding only the first N rows is not a substitute for the file."""
    monkeypatch.setattr(workbook, "MAX_ROWS_PER_SHEET", 1)

    result = report.build(crawl, "Example", consolidate=True)["consolidation"]

    kept = {k["file"] for k in result["kept"]}
    assert kept, "truncated exports must be kept"
    assert kept <= _csv_names(crawl), "every kept file is still on disk"
    assert not (kept & {d["file"] for d in result["deleted"]})


def test_a_failed_workbook_deletes_nothing(crawl: Path):  # noqa: F811
    before = _csv_names(crawl)
    carried = {name: {"sheet": "Nope", "rows_written": 1, "rows_total": 1,
                      "columns": 1, "complete": True} for name in before}

    result = consolidate.run(crawl, crawl / "does-not-exist.xlsx", carried)

    assert not result["ok"]
    assert result["deleted"] == []
    assert _csv_names(crawl) == before


def test_build_report_keeps_exports_by_default(crawl: Path):  # noqa: F811
    before = _csv_names(crawl)
    result = report.build(crawl, "Example")
    assert "consolidation" not in result
    assert _csv_names(crawl) == before
