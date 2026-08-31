"""Tests that run without Screaming Frog installed.

Everything here works off synthetic export CSVs, so the parsing, scoring and
derived analysis are verified on a machine that has never seen the SEO Spider.
"""

import csv
import json
from pathlib import Path

import pytest

from screamingfrog_audit_mcp import analysis, report

INTERNAL_COLUMNS = [
    "Address", "Content Type", "Status Code", "Indexability",
    "Indexability Status", "Title 1", "H1-1", "Word Count", "Crawl Depth",
    "Link Score", "Unique Inlinks", "Response Time", "Size (Bytes)",
]


def _write(path: Path, columns: list[str], rows: list[dict]) -> None:
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=columns)
        w.writeheader()
        for r in rows:
            w.writerow({c: r.get(c, "") for c in columns})


@pytest.fixture
def crawl(tmp_path: Path) -> Path:
    rows = [
        # a healthy top-level page
        {"Address": "https://example.com/", "Content Type": "text/html",
         "Status Code": "200", "Indexability": "Indexable", "Title 1": "Home",
         "H1-1": "Home", "Word Count": "900", "Crawl Depth": "0",
         "Link Score": "100", "Unique Inlinks": "12", "Response Time": "0.21",
         "Size (Bytes)": "45000"},
        # duplicate title AND h1 with the next one
        {"Address": "https://example.com/a", "Content Type": "text/html",
         "Status Code": "200", "Indexability": "Indexable", "Title 1": "Services",
         "H1-1": "Services", "Word Count": "120", "Crawl Depth": "2",
         "Link Score": "40", "Unique Inlinks": "3", "Response Time": "0.4",
         "Size (Bytes)": "30000"},
        {"Address": "https://example.com/b", "Content Type": "text/html",
         "Status Code": "200", "Indexability": "Indexable", "Title 1": "Services",
         "H1-1": "Services", "Word Count": "80", "Crawl Depth": "5",
         "Link Score": "3", "Unique Inlinks": "0", "Response Time": "2.5",
         "Size (Bytes)": "3000000"},
        # non-indexable, and not in the sitemap
        {"Address": "https://example.com/c", "Content Type": "text/html",
         "Status Code": "200", "Indexability": "Non-Indexable",
         "Indexability Status": "Noindex", "Title 1": "Hidden", "H1-1": "Hidden",
         "Word Count": "1500", "Crawl Depth": "3", "Link Score": "20",
         "Unique Inlinks": "1", "Response Time": "0.3", "Size (Bytes)": "50000"},
    ]
    _write(tmp_path / "internal_all.csv", INTERNAL_COLUMNS, rows)
    _write(tmp_path / "internal_html.csv", INTERNAL_COLUMNS, rows)

    _write(tmp_path / "sitemaps_all.csv", ["Address"], [
        {"Address": "https://example.com/"},
        {"Address": "https://example.com/a"},
        {"Address": "https://example.com/c"},   # listed but non-indexable
    ])

    _write(tmp_path / "issues_overview_report.csv",
           ["Issue Name", "Issue Type", "Issue Priority", "URLs", "% of Total",
            "Description", "How To Fix"],
           [
               {"Issue Name": "Low priority thing", "Issue Type": "Opportunity",
                "Issue Priority": "Low", "URLs": "9", "% of Total": "20.0",
                "Description": "d", "How To Fix": "f"},
               {"Issue Name": "Broken pages", "Issue Type": "Issue",
                "Issue Priority": "High", "URLs": "3", "% of Total": "10.0",
                "Description": "d", "How To Fix": "Fix the links."},
               {"Issue Name": "Duplicate titles", "Issue Type": "Warning",
                "Issue Priority": "Medium", "URLs": "2", "% of Total": "5.0",
                "Description": "d", "How To Fix": "Rewrite them."},
           ])
    return tmp_path


# ── issue register ───────────────────────────────────────────────────────────

def test_issues_rank_high_first_regardless_of_file_order(crawl):
    issues = analysis.read_issues(crawl)
    assert [i["priority"] for i in issues] == ["High", "Medium", "Low"]
    assert issues[0]["issue"] == "Broken pages"


def test_health_score_weights_priority_over_volume():
    """Forty cosmetic opportunities must not outrank one real defect."""
    one_defect = analysis.health_score(
        [{"priority": "High", "type": "Issue"}])
    many_cosmetic = analysis.health_score(
        [{"priority": "Low", "type": "Opportunity"}] * 20)
    assert one_defect["score"] < many_cosmetic["score"]
    assert "formula" in one_defect


def test_health_score_is_clamped():
    assert analysis.health_score([]) ["score"] == 100
    flood = [{"priority": "High", "type": "Issue"}] * 500
    assert analysis.health_score(flood)["score"] == 0


def test_summarize_writes_and_counts(crawl):
    s = analysis.summarize(crawl, "https://example.com")
    assert (crawl / "audit-summary.json").exists()
    assert s["stats"]["urls"] == 4
    assert s["stats"]["indexable"] == 3
    assert s["stats"]["non_indexable"] == 1
    assert s["counts"] == {"high": 1, "medium": 1, "low": 1, "total_types": 3}


# ── derived analysis ─────────────────────────────────────────────────────────

def test_depth_flags_only_pages_past_three(crawl):
    d = analysis.full_analysis(crawl)["depth"]
    assert d["max_depth"] == 5
    assert d["pages_deeper_than_3"] == 1
    assert d["deepest"][0]["url"].endswith("/b")


def test_link_equity_finds_the_starved_page(crawl):
    le = analysis.full_analysis(crawl)["link_equity"]
    assert le["no_internal_inlinks"] == 1
    assert le["least_linked"][0]["unique_inlinks"] == 0


def test_sitemap_reconciles_both_directions(crawl):
    sm = analysis.full_analysis(crawl)["sitemap"]
    assert sm["available"] is True
    # /c is listed but non-indexable
    assert sm["in_sitemap_but_not_indexable"] == 1
    # /b is indexable but absent from the sitemap
    assert "https://example.com/b" in sm["examples_missing"]


def test_sitemap_degrades_when_absent(tmp_path):
    _write(tmp_path / "internal_all.csv", INTERNAL_COLUMNS, [])
    sm = analysis.full_analysis(tmp_path)["sitemap"]
    assert sm["available"] is False


def test_content_thin_pages(crawl):
    ct = analysis.full_analysis(crawl)["content"]
    assert ct["thin_pages"] == 2          # 120 and 80 words
    assert ct["thinnest"][0]["words"] == 80


def test_performance_outliers(crawl):
    perf = analysis.full_analysis(crawl)["performance"]
    assert perf["slow_pages"] == 1
    assert perf["heavy_pages"] == 1
    assert perf["slowest"][0]["seconds"] == 2.5


def test_indexability_reasons(crawl):
    ix = analysis.full_analysis(crawl)["indexability"]
    assert ix["indexable"] == 3
    assert ix["reasons"] == {"Noindex": 1}


def test_duplication_groups(crawl):
    dup = analysis.full_analysis(crawl)["duplication"]
    assert dup["duplicate_title_groups"] == 1
    assert dup["worst_titles"][0]["count"] == 2


def test_analysis_survives_missing_and_junk_values(tmp_path):
    """A real export has blanks and non-numeric cells; none may raise."""
    _write(tmp_path / "internal_all.csv", INTERNAL_COLUMNS, [
        {"Address": "https://example.com/", "Content Type": "text/html",
         "Word Count": "", "Crawl Depth": "n/a", "Unique Inlinks": None,
         "Response Time": "", "Size (Bytes)": "1,024"},
    ])
    result = analysis.full_analysis(tmp_path)
    assert set(result) == {"depth", "link_equity", "sitemap", "content",
                           "performance", "indexability", "duplication"}


# ── report ───────────────────────────────────────────────────────────────────

def test_report_writes_all_three_files(crawl):
    out = report.build(crawl, "Example Co")
    for key in ("markdown", "html", "analysis_json"):
        assert Path(out[key]).exists()
    assert out["health"]["score"] <= 100


def test_report_html_escapes_hostile_content(tmp_path):
    """Crawl data is attacker-controllable; it must never become markup."""
    _write(tmp_path / "internal_all.csv", INTERNAL_COLUMNS, [
        {"Address": "https://example.com/<script>alert(1)</script>",
         "Content Type": "text/html", "Indexability": "Indexable",
         "Title 1": "<img src=x onerror=alert(1)>", "H1-1": "x",
         "Word Count": "50", "Crawl Depth": "1", "Unique Inlinks": "1",
         "Response Time": "0.1", "Size (Bytes)": "100"},
    ])
    _write(tmp_path / "issues_overview_report.csv",
           ["Issue Name", "Issue Type", "Issue Priority", "URLs", "% of Total",
            "Description", "How To Fix"],
           [{"Issue Name": "<b>bold</b>", "Issue Type": "Issue",
             "Issue Priority": "High", "URLs": "1", "% of Total": "100",
             "Description": "d", "How To Fix": "<script>x</script>"}])
    out = report.build(tmp_path, "Example")
    page = Path(out["html"]).read_text(encoding="utf-8")
    assert "<script>alert(1)</script>" not in page
    assert "onerror=alert(1)" not in page
    assert "&lt;script&gt;" in page


def test_report_rebuilds_stale_summary_without_health(crawl):
    """A summary written by an older version lacks 'health'; it must be
    regenerated rather than crashing the report."""
    stale = {"site": "https://example.com", "stats": {}, "counts": {}, "issues": []}
    (crawl / "audit-summary.json").write_text(json.dumps(stale), encoding="utf-8")
    out = report.build(crawl)
    assert out["health"]["score"] > 0


# ── expert-mode tier logic ───────────────────────────────────────────────────

from screamingfrog_audit_mcp import crawler  # noqa: E402


def test_expert_mode_excludes_api_groups_on_free():
    excluded = crawler._expert_excluded(licensed=False, has_config=False)
    assert "Search Console" in excluded
    assert "PageSpeed" in excluded
    assert "Custom Extraction" in excluded
    assert "UNDEF" in excluded


def test_expert_mode_includes_api_groups_once_licensed():
    excluded = crawler._expert_excluded(licensed=True, has_config=False)
    assert "Search Console" not in excluded
    assert "Change Detection" not in excluded
    # still excluded: these only fill in from a config file
    assert "Custom Search" in excluded


def test_config_unlocks_the_custom_groups():
    excluded = crawler._expert_excluded(licensed=True, has_config=True)
    assert not (excluded - {"UNDEF"})


def test_export_args_passes_config_through(tmp_path, monkeypatch):
    monkeypatch.setattr(crawler, "accepted_names", lambda kind: set())
    cfg = tmp_path / "my.seospiderconfig"
    cfg.write_text("x", encoding="utf-8")
    args = crawler.export_args(tmp_path, everything=False, config=str(cfg))
    assert "--config" in args
    assert str(cfg) in args


def test_export_args_omits_config_when_absent(tmp_path, monkeypatch):
    monkeypatch.setattr(crawler, "accepted_names", lambda kind: set())
    assert "--config" not in crawler.export_args(tmp_path)


def test_unknown_filter_names_are_dropped_not_passed_through(tmp_path, monkeypatch):
    """A name the binary does not know aborts the whole crawl, so it must be
    filtered out rather than forwarded."""
    monkeypatch.setattr(crawler, "accepted_names",
                        lambda kind: {"Internal:All", "Crawl Overview"})
    args = crawler.export_args(tmp_path)
    tabs = args[args.index("--export-tabs") + 1]
    reports = args[args.index("--save-report") + 1]
    assert tabs == "Internal:All"
    assert reports == "Crawl Overview"


# ── MCP SDK compatibility ────────────────────────────────────────────────────

def test_server_imports_and_registers_every_tool():
    """Guards the SDK rename: mcp 2.0 moved FastMCP to MCPServer, and a fresh
    `pip install mcp` now resolves to 2.x. Importing under whichever major is
    present must work, or every new install crashes on startup."""
    from screamingfrog_audit_mcp import server as srv

    assert srv.server is not None
    for name in ("check_install", "available_filters", "start_crawl",
                 "crawl_status", "cancel_crawl", "list_crawls", "get_issues",
                 "get_analysis", "list_exports", "read_export", "build_report"):
        assert callable(getattr(srv, name)), f"{name} is not exposed"


def test_server_module_does_not_shadow_the_mcp_package():
    """`mcp = FastMCP(...)` would rebind the package name inside the module."""
    import mcp as sdk

    from screamingfrog_audit_mcp import server as srv
    assert getattr(srv, "mcp", sdk) is sdk


# ── Windows process safety ───────────────────────────────────────────────────

def test_windows_liveness_never_calls_os_kill(monkeypatch, tmp_path):
    """On Windows, os.kill with any signal but CTRL_C/CTRL_BREAK routes to
    TerminateProcess. Using it as a liveness probe would kill the crawl and
    then report it finished."""
    from screamingfrog_audit_mcp import jobs as J

    monkeypatch.setattr(J.sys, "platform", "win32")

    def boom(*a, **k):
        raise AssertionError("os.kill must never be called on Windows")

    monkeypatch.setattr(J.os, "kill", boom)
    monkeypatch.setattr(J.subprocess, "run",
                        lambda *a, **k: type("R", (), {"stdout": "runner.exe 4321 Console"})())

    j = J.Jobs(tmp_path)
    assert j.running({"job_id": "absent", "pid": 4321}) is True


def test_windows_liveness_reports_dead_when_pid_absent(monkeypatch, tmp_path):
    from screamingfrog_audit_mcp import jobs as J

    monkeypatch.setattr(J.sys, "platform", "win32")
    monkeypatch.setattr(J.subprocess, "run",
                        lambda *a, **k: type("R", (), {
                            "stdout": "INFO: No tasks are running which match the criteria."})())
    j = J.Jobs(tmp_path)
    assert j.running({"job_id": "absent", "pid": 4321}) is False


def test_windows_cancel_uses_taskkill_not_killpg(monkeypatch, tmp_path):
    """os.killpg and os.getpgid do not exist on Windows at all."""
    from screamingfrog_audit_mcp import jobs as J

    monkeypatch.setattr(J.sys, "platform", "win32")
    monkeypatch.setattr(J.os, "kill", lambda *a, **k: (_ for _ in ()).throw(
        AssertionError("must not signal on Windows")))
    seen = {}
    monkeypatch.setattr(J.subprocess, "run",
                        lambda cmd, **k: seen.setdefault("cmd", cmd))

    j = J.Jobs(tmp_path)
    job = {"job_id": "x", "pid": 4321, "output_dir": str(tmp_path)}
    j.cancel(job)
    assert seen["cmd"][0] == "taskkill"
    assert "/F" in seen["cmd"]
    assert job["state"] == "cancelled"


def test_posix_liveness_treats_a_zombie_as_finished(monkeypatch, tmp_path):
    """The original bug: a reaped-less child polls 'running' forever."""
    from screamingfrog_audit_mcp import jobs as J

    monkeypatch.setattr(J.sys, "platform", "darwin")
    monkeypatch.setattr(J.os, "kill", lambda *a, **k: None)
    monkeypatch.setattr(J.subprocess, "run",
                        lambda *a, **k: type("R", (), {"stdout": "Z+\n"})())
    j = J.Jobs(tmp_path)
    assert j.running({"job_id": "absent", "pid": 4321}) is False


def test_no_subprocess_decodes_with_the_locale_encoding():
    """text=True without an explicit encoding decodes as cp1252 on Windows, so
    the first non-ASCII byte raises UnicodeDecodeError."""
    import re
    from pathlib import Path

    import screamingfrog_audit_mcp

    src = Path(screamingfrog_audit_mcp.__file__).parent
    offenders = []
    for f in src.glob("*.py"):
        text = f.read_text(encoding="utf-8")
        for call in re.findall(r"subprocess\.run\((?:[^()]|\([^()]*\))*\)", text):
            if "text=True" in call and "encoding=" not in call:
                offenders.append(f"{f.name}: {' '.join(call.split())[:90]}")
    assert not offenders, "unpinned decode: " + "; ".join(offenders)


# ── doctor ───────────────────────────────────────────────────────────────────

def test_doctor_fails_when_the_spider_is_missing(monkeypatch, capsys, tmp_path):
    from screamingfrog_audit_mcp import doctor as doc

    monkeypatch.setenv("SCREAMING_FROG_PATH", str(tmp_path / "nope"))
    monkeypatch.setenv("SF_MCP_AUDIT_DIR", str(tmp_path / "audits"))
    assert doc.run() == 1
    out = capsys.readouterr().out
    assert "[FAIL]" in out
    assert "screamingfrog.co.uk" in out          # tells them where to get it
    assert "Looked in" in out                    # and where it searched


def test_doctor_passes_and_prints_a_usable_config(monkeypatch, capsys, tmp_path):
    from screamingfrog_audit_mcp import doctor as doc

    fake = tmp_path / "spider"
    fake.write_text("#!/bin/sh\n", encoding="utf-8")
    monkeypatch.setenv("SCREAMING_FROG_PATH", str(fake))
    monkeypatch.setenv("SF_MCP_AUDIT_DIR", str(tmp_path / "audits"))
    monkeypatch.setattr(doc, "_check_binary_runs",
                        lambda: (doc.PASS, "Binary responds: 900 export filters", ""))
    assert doc.run() == 0
    out = capsys.readouterr().out
    assert "All checks passed" in out
    assert '"mcpServers"' in out
    assert "[FAIL]" not in out


def test_doctor_treats_the_free_tier_as_a_warning_not_a_failure(monkeypatch, tmp_path):
    """The free tier is fully supported; flagging it as a failure would send
    people off to buy a licence they do not need."""
    from screamingfrog_audit_mcp import doctor as doc

    monkeypatch.setattr(doc, "is_licensed", lambda: False)
    status, _, hint = doc._check_licence()
    assert status == doc.WARN
    assert "fully supported" in hint


def test_version_is_single_sourced_from_distribution_metadata():
    """The version was hardcoded in __init__.py once and drifted: the package
    shipped as 1.0.1 while --doctor still reported 1.0.0. pyproject.toml must
    stay the only place a version number lives."""
    import importlib.metadata as md
    import re
    from pathlib import Path

    import screamingfrog_audit_mcp as pkg

    assert pkg.__version__ == md.version("screamingfrog-audit-mcp")

    # It must be derived, not restated. The fallback literal for an
    # uninstalled source tree is allowed; a release number is not.
    init = Path(pkg.__file__).read_text(encoding="utf-8")
    assert "importlib.metadata" in init, "version is no longer derived from metadata"
    hardcoded = re.findall(r'__version__\s*=\s*"([^"]+)"', init)
    assert all(v.endswith("+source") for v in hardcoded), (
        f"a release version is hardcoded in __init__.py: {hardcoded}")


# ── domain allowlist ─────────────────────────────────────────────────────────

def test_allowlist_permits_the_domain_and_its_subdomains(monkeypatch):
    from screamingfrog_audit_mcp import server as srv

    monkeypatch.setenv("SF_ALLOWED_DOMAINS", "example.com, acme.co.uk")
    srv._check_allowed("https://example.com/page")
    srv._check_allowed("https://www.example.com/")
    srv._check_allowed("https://shop.acme.co.uk/x")


def test_allowlist_blocks_everything_else(monkeypatch):
    """Without this, a confused or prompt-injected agent can point the crawler
    at internal hosts or unrelated third parties."""
    from screamingfrog_audit_mcp import server as srv

    monkeypatch.setenv("SF_ALLOWED_DOMAINS", "example.com")
    for bad in ("https://evil.com/", "http://169.254.169.254/latest/meta-data/",
                "https://notexample.com/", "https://example.com.evil.com/"):
        with pytest.raises(ValueError, match="not in SF_ALLOWED_DOMAINS"):
            srv._check_allowed(bad)


def test_no_allowlist_means_no_restriction(monkeypatch):
    from screamingfrog_audit_mcp import server as srv

    monkeypatch.delenv("SF_ALLOWED_DOMAINS", raising=False)
    srv._check_allowed("https://anything.example/")


# ── filter modes ─────────────────────────────────────────────────────────────

def test_filter_modes(monkeypatch):
    from screamingfrog_audit_mcp import server as srv

    assert srv._matcher("FOO", "contains")("a foo b") is True
    assert srv._matcher("foo", "exact")("FOO") is True
    assert srv._matcher("foo", "exact")("foo bar") is False
    assert srv._matcher(r"^\d{3}$", "regex")("404") is True
    assert srv._matcher(r"^\d{3}$", "regex")("4041") is False
    assert srv._matcher("", "contains")("anything") is True
    with pytest.raises(ValueError, match="Invalid regex"):
        srv._matcher("[unclosed", "regex")
    with pytest.raises(ValueError, match="mode must be"):
        srv._matcher("x", "fuzzy")


def test_read_export_filters_by_named_column(crawl, monkeypatch, tmp_path):
    from screamingfrog_audit_mcp import server as srv

    monkeypatch.setattr(srv, "AUDIT_ROOT", crawl.parent)
    out = srv.read_export(crawl=crawl.name, export="internal_all",
                          column="Indexability", contains="Non-Indexable",
                          mode="exact", columns="Address,Indexability")
    assert out["total_matching_rows"] == 1
    assert out["rows"][0]["Address"].endswith("/c")


def test_read_export_rejects_an_unknown_filter_column(crawl, monkeypatch):
    from screamingfrog_audit_mcp import server as srv

    monkeypatch.setattr(srv, "AUDIT_ROOT", crawl.parent)
    with pytest.raises(ValueError, match="No column"):
        srv.read_export(crawl=crawl.name, export="internal_all", column="Nope",
                        contains="x")


# ── aggregation ──────────────────────────────────────────────────────────────

def test_aggregate_counts_without_returning_rows(crawl, monkeypatch):
    """The point: answer 'how many, broken down by' in one small response."""
    from screamingfrog_audit_mcp import server as srv

    monkeypatch.setattr(srv, "AUDIT_ROOT", crawl.parent)
    out = srv.aggregate_export(crawl=crawl.name, export="internal_all",
                               group_by="Indexability")
    assert out["rows_matched"] == 4
    by = {g["value"]: g["rows"] for g in out["groups"]}
    assert by == {"Indexable": 3, "Non-Indexable": 1}
    assert "rows" not in out                      # never returns the rows


def test_aggregate_summarises_a_numeric_metric(crawl, monkeypatch):
    from screamingfrog_audit_mcp import server as srv

    monkeypatch.setattr(srv, "AUDIT_ROOT", crawl.parent)
    out = srv.aggregate_export(crawl=crawl.name, export="internal_all",
                               group_by="Indexability", metric="Word Count")
    indexable = next(g for g in out["groups"] if g["value"] == "Indexable")
    assert indexable["metric"]["sum"] == 1100.0    # 900 + 120 + 80
    assert indexable["metric"]["max"] == 900.0


def test_aggregate_rejects_unknown_columns(crawl, monkeypatch):
    from screamingfrog_audit_mcp import server as srv

    monkeypatch.setattr(srv, "AUDIT_ROOT", crawl.parent)
    with pytest.raises(ValueError, match="No column 'Nope' for group_by"):
        srv.aggregate_export(crawl=crawl.name, export="internal_all", group_by="Nope")


# ── storage ──────────────────────────────────────────────────────────────────

def test_delete_crawl_requires_confirmation(crawl, monkeypatch):
    from screamingfrog_audit_mcp import server as srv

    monkeypatch.setattr(srv, "AUDIT_ROOT", crawl.parent)
    out = srv.delete_crawl(crawl=crawl.name)
    assert out["deleted"] is False
    assert crawl.exists(), "must not delete without confirm"

    out = srv.delete_crawl(crawl=crawl.name, confirm=True)
    assert out["deleted"] is True
    assert not crawl.exists()


def test_delete_crawl_refuses_paths_outside_the_audit_root(tmp_path, monkeypatch):
    """A crafted absolute path must never delete arbitrary directories."""
    from screamingfrog_audit_mcp import server as srv

    root = tmp_path / "audits"
    (root / "a-crawl").mkdir(parents=True)
    monkeypatch.setattr(srv, "AUDIT_ROOT", root)

    outside = tmp_path / "precious"
    outside.mkdir()
    (outside / "keep.txt").write_text("x", encoding="utf-8")

    with pytest.raises(ValueError, match="outside the audit root"):
        srv.delete_crawl(crawl=str(outside), confirm=True)
    assert (outside / "keep.txt").exists()

    with pytest.raises(ValueError, match="audit root itself"):
        srv.delete_crawl(crawl=str(root), confirm=True)
    assert root.exists()


def test_storage_summary_reports_sizes(tmp_path, monkeypatch):
    from screamingfrog_audit_mcp import server as srv

    root = tmp_path / "audits"
    folder = root / "site-2026-08-31"
    folder.mkdir(parents=True)
    (folder / "internal_all.csv").write_text("x" * 2048, encoding="utf-8")
    (root / ".jobs").mkdir()                     # hidden, must be skipped
    monkeypatch.setattr(srv, "AUDIT_ROOT", root)

    out = srv.storage_summary()
    assert out["crawls"] == 1
    assert out["largest"][0]["crawl"] == "site-2026-08-31"
    assert out["largest"][0]["files"] == 1
    assert out["total_mb"] >= 0


# ── branding and the master workbook ─────────────────────────────────────────

def test_every_output_carries_the_credit(crawl):
    """The whole point of the branding: no generated artefact loses it."""
    from screamingfrog_audit_mcp import branding as B

    out = report.build(crawl, "Example Co")
    md = Path(out["markdown"]).read_text(encoding="utf-8")
    html = Path(out["html"]).read_text(encoding="utf-8")

    for text in (md, html):
        assert B.AUTHOR in text
        assert B.AUTHOR_SITE_SHORT in text
    assert B.AUTHOR_SITE in html          # a real, clickable link
    assert B.STUDIO in html


def test_html_uses_the_brand_palette(crawl):
    from screamingfrog_audit_mcp import branding as B

    html = Path(report.build(crawl, "Example Co")["html"]).read_text(encoding="utf-8")
    assert B.css(B.PURPLE) in html
    assert B.css(B.CREAM) in html
    assert 'class="masthead"' in html


def test_workbook_has_a_sheet_for_every_export(crawl):
    from openpyxl import load_workbook

    out = report.build(crawl, "Example Co")
    wb = load_workbook(out["workbook"])
    # Summary + Issue register + Analysis + Data index + one per non-empty export
    exports = [p for p in crawl.glob("*.csv")]
    assert out["data_tables"] == len(exports)
    for fixed in ("Summary", "Issue register", "Analysis", "Data index"):
        assert fixed in wb.sheetnames


def test_workbook_highlights_priority_as_colour(crawl):
    """A reader must see severity without reading the word."""
    from openpyxl import load_workbook

    from screamingfrog_audit_mcp import branding as B

    wb = load_workbook(report.build(crawl, "Example Co")["workbook"])
    ws = wb["Issue register"]
    assert ws["A5"].fill.fgColor.rgb.endswith(B.PURPLE)     # header
    chips = {ws.cell(row=r, column=1).value:
             ws.cell(row=r, column=1).fill.fgColor.rgb for r in range(6, 9)}
    assert any(v.endswith(B.RED) for v in chips.values())
    assert any(v.endswith(B.AMBER) for v in chips.values())


def test_workbook_sheets_are_navigable(crawl):
    """Frozen header, autofilter and a coloured tab on every data sheet."""
    from openpyxl import load_workbook

    wb = load_workbook(report.build(crawl, "Example Co")["workbook"])
    ws = wb["internal all"]
    assert ws.freeze_panes == "A2"
    assert ws.auto_filter.ref is not None
    assert ws.sheet_properties.tabColor is not None
    assert ws.sheet_view.showGridLines is False


def test_sheet_names_are_excel_safe():
    from screamingfrog_audit_mcp.workbook import _safe_sheet_name

    taken = set()
    long = _safe_sheet_name("a" * 60, taken)
    assert len(long) <= 31
    assert _safe_sheet_name("has/bad:chars*", set()) == "has-bad-chars-"
    # duplicates must not collide, Excel refuses two sheets with one name
    t = set()
    first = _safe_sheet_name("internal_all", t)
    second = _safe_sheet_name("internal_all", t)
    assert first != second


# ── version compatibility (Screaming Frog 19.8 regression) ───────────────────

# What an older build advertises. Note: no --skip-empty. Passing it to 19.8
# aborts the whole crawl with UnrecognizedOptionException, which is exactly
# what a user hit on Windows.
_OLD_HELP_FLAGS = {
    "--headless", "--crawl", "--crawl-list", "--output-folder", "--overwrite",
    "--export-tabs", "--save-report", "--config", "--help",
}


def _fake_build(monkeypatch, flags):
    from screamingfrog_audit_mcp import crawler as c

    monkeypatch.setattr(c, "_FLAG_CACHE", set(flags))
    monkeypatch.setattr(c, "accepted_names", lambda kind: set())
    return c


def test_skip_empty_is_never_sent_to_a_build_without_it(tmp_path, monkeypatch):
    """The 19.8 crash: one unrecognised flag aborts the entire crawl."""
    c = _fake_build(monkeypatch, _OLD_HELP_FLAGS)
    args = c.export_args(tmp_path)
    assert "--skip-empty" not in args
    assert "--overwrite" in args          # this one IS supported, so keep it
    assert "--headless" in args


def test_skip_empty_is_sent_when_the_build_supports_it(tmp_path, monkeypatch):
    c = _fake_build(monkeypatch, _OLD_HELP_FLAGS | {"--skip-empty"})
    assert "--skip-empty" in c.export_args(tmp_path)


def test_unknown_option_set_means_omit_every_optional_flag(tmp_path, monkeypatch):
    """If --help cannot be read we must not guess: an unknown flag is fatal,
    a missing optional flag only costs a few empty CSVs."""
    c = _fake_build(monkeypatch, set())
    args = c.export_args(tmp_path)
    for flag in c.OPTIONAL_FLAGS:
        assert flag not in args
    assert "--headless" in args and "--output-folder" in args


def test_config_is_refused_rather_than_crashing_the_crawl(tmp_path, monkeypatch):
    c = _fake_build(monkeypatch, _OLD_HELP_FLAGS - {"--config"})
    cfg = tmp_path / "x.seospiderconfig"
    cfg.write_text("x", encoding="utf-8")
    with pytest.raises(RuntimeError, match="does not support --config"):
        c.export_args(tmp_path, config=str(cfg))


def test_missing_essential_flags_are_reported(monkeypatch):
    c = _fake_build(monkeypatch, {"--help"})
    gone = c.missing_essentials()
    assert "--headless" in gone and "--crawl" in gone


def test_no_essentials_reported_when_help_is_unreadable(monkeypatch):
    """Empty means 'unknown', not 'everything is missing'."""
    c = _fake_build(monkeypatch, set())
    assert c.missing_essentials() == []


def test_supported_flags_parses_real_help_output(monkeypatch):
    from screamingfrog_audit_mcp import crawler as c

    help_text = (
        "Usage: ScreamingFrogSEOSpider options\n\n"
        "Options:\n"
        "    --crawl <url>\n         Start crawling the supplied URL\n"
        "    --headless\n         Run in headless mode\n"
        "    --export-tabs <tabs>\n         Supply a comma separated list\n"
    )
    monkeypatch.setattr(c, "_FLAG_CACHE", None)
    monkeypatch.setattr(c, "find_binary", lambda: Path("/fake/sf"))
    monkeypatch.setattr(c.subprocess, "run",
                        lambda *a, **k: type("R", (), {"stdout": help_text})())
    flags = c.supported_flags(refresh=True)
    assert flags == {"--crawl", "--headless", "--export-tabs"}


def test_doctor_flags_check_fails_on_a_broken_build(monkeypatch):
    from screamingfrog_audit_mcp import crawler as c
    from screamingfrog_audit_mcp import doctor as doc

    monkeypatch.setattr(c, "_FLAG_CACHE", {"--help"})
    status, message, _ = doc._check_flags()
    assert status == doc.FAIL
    assert "Missing essential options" in message
